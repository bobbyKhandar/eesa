import os
from openpyxl import load_workbook, Workbook

def merge_excels_with_branch(input_dir, output_file):
    merged_wb = Workbook()
    merged_ws = merged_wb.active
    merged_ws.title = "MergedQuestions"

    header_written = False

    for file in os.listdir(input_dir):
        if not file.endswith(".xlsx"):
            continue

        file_path = os.path.join(input_dir, file)
        branch_name = os.path.splitext(file)[0]  # e.g., 'IT.xlsx' → 'IT'

        try:
            wb = load_workbook(file_path)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                continue

            # Write header once, with 'Branch' column added
            if not header_written:
                merged_ws.append(rows[0] + ("branch",))
                header_written = True

            # Append data rows with branch column
            for row in rows[1:]:
                merged_ws.append(row + (branch_name,))

            print(f"✅ Appended: {file} ({len(rows)-1} questions)")

        except Exception as e:
            print(f"❌ Failed: {file} → {e}")   

    merged_wb.save(output_file)
    print(f"\n✅ Merged all into: {output_file}")

# === USAGE ===
input_dir = "C:/project/miniproject/outputs/excel_exports"
output_file = "C:/project/miniproject/outputs/merged_questions_with_branch.xlsx"
merge_excels_with_branch(input_dir, output_file)
